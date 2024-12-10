from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from .models import PizzaOrder
from django.template.loader import render_to_string
import tempfile

def order_pizza(request):
    if request.method == "POST":
        size = request.POST.get("size")
        pepperoni = request.POST.get("pepperoni") == "Y"
        extra_cheese = request.POST.get("extra_cheese") == "Y"

        # Calculate the bill
        base_prices = {"Small": 100, "Medium": 150, "Large": 200}
        bill = base_prices.get(size, 0)
        if pepperoni:
            bill += 50  # Add pepperoni price
        if extra_cheese:
            bill += 30  # Add extra cheese price

        # Save the order with the bill
        order = PizzaOrder(size=size, pepperoni=pepperoni, extra_cheese=extra_cheese, bill=bill)
        order.save()

        # Return JSON response with order details and order ID for redirect
        response_data = {
            "size": order.size,
            "pepperoni": "Yes" if order.pepperoni else "No",
            "extra_cheese": "Yes" if order.extra_cheese else "No",
            "bill": float(order.bill),
            "order_id": order.id,  # Include order ID
        }

        return JsonResponse(response_data)

    return render(request, "order_form.html")




from django.http import JsonResponse

def get_pizza_price(request):
    if request.method == "POST":
        import json
        data = json.loads(request.body)

        # Extract options
        size = data.get("size", "Small")
        pepperoni = data.get("pepperoni") == "Y"
        extra_cheese = data.get("extra_cheese") == "Y"

        # Calculate price
        base_prices = {"Small": 100, "Medium": 150, "Large": 200}
        price = base_prices.get(size, 0)
        if pepperoni:
            price += 50  # Add pepperoni price
        if extra_cheese:
            price += 30  # Add extra cheese price

        # Return price as JSON
        return JsonResponse({"price": price})

    return JsonResponse({"error": "Invalid request method"}, status=400)



def confirmation_page(request, order_id):
    try:
        order = PizzaOrder.objects.get(pk=order_id)
    except PizzaOrder.DoesNotExist:
        return HttpResponse("Order not found", status=404)

    return render(request, "confirmation.html", {"order": order})




from django.http import HttpResponse
from reportlab.pdfgen import canvas
from io import BytesIO
from .models import PizzaOrder

def download_invoice(request, order_id):
    try:
        order = PizzaOrder.objects.get(pk=order_id)
    except PizzaOrder.DoesNotExist:
        return HttpResponse("Order not found", status=404)

    # Create a BytesIO buffer to store the PDF
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)

    # Set the title
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(100, 800, "Pizza Order Invoice")

    # Add order details
    pdf.setFont("Helvetica", 12)
    pdf.drawString(100, 750, f"Order ID: {order.id}")
    pdf.drawString(100, 730, f"Size: {order.size}")
    pdf.drawString(100, 710, f"Pepperoni: {'Yes' if order.pepperoni else 'No'}")
    pdf.drawString(100, 690, f"Extra Cheese: {'Yes' if order.extra_cheese else 'No'}")
    pdf.drawString(100, 670, f"Bill: Rs {order.bill:.2f}")

    # Finalize the PDF
    pdf.save()

    # Get the PDF from the buffer
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f"attachment; filename=Invoice_{order.id}.pdf"

    return response


import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import PizzaOrder

def download_invoice_excel(request, order_id):
    try:
        order = PizzaOrder.objects.get(pk=order_id)
    except PizzaOrder.DoesNotExist:
        return HttpResponse("Order not found", status=404)

    # Create a workbook and worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"

    # Add title with a larger, bold font
    sheet["A1"] = "Pizza Order Invoice"
    sheet["A1"].font = Font(bold=True, size=16, color="0000FF")  # Blue and bold
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Add a line break
    sheet.append([])

    # Add headers with background color
    header = ["Order ID", "Size", "Pepperoni", "Extra Cheese", "Bill"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue fill
    for col_num, header_name in enumerate(header, 1):
        cell = sheet.cell(row=3, column=col_num, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add order details
    order_data = [
        order.id,
        order.size,
        "Yes" if order.pepperoni else "No",
        "Yes" if order.extra_cheese else "No",
        f"Rs {order.bill:.2f}"
    ]
    
    for col_num, data in enumerate(order_data, 1):
        cell = sheet.cell(row=4, column=col_num, value=data)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Apply borders to cells
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border

    # Set column width for better readability
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 10

    # Save workbook to an in-memory buffer
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename=Invoice_{order.id}.xlsx'
    workbook.save(response)

    return response



